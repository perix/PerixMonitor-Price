"""
exports.py — Export aggiuntivi in formato Excel (xlsx).

- build_report_workbook: esiti della lettura in due fogli
  ("Recuperati" e "Non Recuperati");
- build_executed_workbook: copia del file di input con tre colonne
  aggiunte: "Price Recovered" (Yes/No), "Trade Site" (borsa usata),
  "Source" (fonte dati).
"""

import io

from openpyxl import Workbook, load_workbook


def build_report_workbook(results) -> bytes:
    """Workbook xlsx con gli esiti: fogli 'Recuperati' e 'Non Recuperati'."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Recuperati"
    ws.append(["ISIN", "Descrizione Asset", "Simbolo", "Borsa",
               "Valuta", "Fonte", "Settimane"])
    for r in results:
        if r.ok:
            ws.append([r.isin, r.description, r.symbol, r.exchange,
                       r.currency, r.source, len(r.prices)])
    for col, width in zip("ABCDEFG", (16, 32, 18, 16, 12, 16, 11)):
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("Non Recuperati")
    ws2.append(["ISIN", "Descrizione Asset", "Motivo"])
    for r in results:
        if not r.ok:
            ws2.append([r.isin, r.description, r.reason])
    for col, width in zip("ABC", (16, 32, 100)):
        ws2.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_executed_workbook(input_bytes: bytes, results) -> bytes:
    """
    Copia del file di input con tre colonne aggiunte in coda:
    'Price Recovered' (Yes/No), 'Trade Site', 'Source'.
    Le righe sono associate ai risultati tramite la colonna ISIN.
    """
    by_isin = {r.isin: r for r in results}

    wb = load_workbook(io.BytesIO(input_bytes))
    ws = wb.active

    # individua la colonna ISIN nell'intestazione (riga 1)
    isin_col = None
    last_col = ws.max_column
    for c in range(1, last_col + 1):
        if str(ws.cell(1, c).value or "").strip().upper() == "ISIN":
            isin_col = c
            break
    if isin_col is None:
        raise ValueError("Colonna 'ISIN' non trovata nel file di input.")

    c_rec, c_site, c_src = last_col + 1, last_col + 2, last_col + 3
    ws.cell(1, c_rec, "Price Recovered")
    ws.cell(1, c_site, "Trade Site")
    ws.cell(1, c_src, "Source")

    for row in range(2, ws.max_row + 1):
        isin = str(ws.cell(row, isin_col).value or "").strip()
        if not isin:
            continue
        r = by_isin.get(isin)
        if r is not None and r.ok:
            ws.cell(row, c_rec, "Yes")
            ws.cell(row, c_site, r.exchange)
            ws.cell(row, c_src, r.source)
        else:
            ws.cell(row, c_rec, "No")

    from openpyxl.utils import get_column_letter
    for c in (c_rec, c_site, c_src):
        ws.column_dimensions[get_column_letter(c)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
