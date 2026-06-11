"""
price_fetcher.py — Logica di recupero prezzi storici settimanali in EUR.

Fonte primaria: investing.com (API interna, vedi investing_fetcher.py).
Fallback: Yahoo Finance (libreria yfinance).

Per ogni ISIN:
  1. tenta investing.com: ricerca ISIN, preferenza Borsa di Milano poi
     area Euro, storico settimanale; preferite le quotazioni EUR native;
  2. se fallisce, tenta Yahoo Finance: ricerca ISIN, varianti di borsa
     (.MI, .DE, ...), preferite le quotazioni EUR native;
  3. se la quotazione disponibile e' in altra valuta, i prezzi vengono
     convertiti in EUR con il cambio storico settimanale (vedi fx.py);
  4. se tutto fallisce, l'asset viene scartato con i motivi.
"""

import datetime as dt
import time
from dataclasses import dataclass, field

import pandas as pd
import yfinance as yf

# Ordine di preferenza dei suffissi Yahoo Finance (borse).
SUFFIX_PREFERENCE = [
    ".MI",                                # Borsa Italiana (Milano)
    ".DE", ".F",                          # Xetra, Francoforte
    ".PA", ".AS", ".BR", ".LS",           # Euronext
    ".MC", ".VI", ".IR",                  # Madrid, Vienna, Dublino
    ".DU", ".MU", ".SG", ".HM", ".HA", ".BE",  # borse regionali tedesche
]

ACCEPTED_TYPES = {"EQUITY", "ETF", "MUTUALFUND", "INDEX", "BOND", "CURRENCY", "FUTURE"}


@dataclass
class FetchResult:
    isin: str
    description: str
    ok: bool
    symbol: str = ""
    exchange: str = ""
    currency: str = ""      # "EUR" oppure "USD→EUR" se convertito
    source: str = ""        # "investing.com" / "Yahoo Finance"
    prices: pd.DataFrame = field(default_factory=pd.DataFrame)  # Data, Prezzo (EUR)
    reason: str = ""


def _to_eur(prices: pd.DataFrame, currency: str, start, end) -> tuple[pd.DataFrame, str]:
    """
    Se serve, converte la serie in EUR. Ritorna (serie, etichetta_valuta).
    Serie vuota se il cambio non e' recuperabile.
    """
    cur = (currency or "").strip()
    if cur.upper() in ("EUR", ""):
        return prices, "EUR"
    from fx import convert_to_eur
    converted = convert_to_eur(prices, cur, start, end)
    if converted.empty:
        return pd.DataFrame(), cur
    return converted, f"{cur}→EUR"


def fetch_asset(isin: str, description: str, start, end, pause: float = 0.6) -> FetchResult:
    """Recupera lo storico settimanale in EUR per un singolo asset."""
    from investing_fetcher import fetch_from_investing

    if isinstance(start, dt.datetime):
        start = start.date()
    if isinstance(end, dt.datetime):
        end = end.date()

    inv_reason = ""
    try:
        df, info, inv_reason = fetch_from_investing(isin, start, end)
        if not df.empty:
            converted, cur_label = _to_eur(df, info.get("currency", "EUR"), start, end)
            if not converted.empty:
                return FetchResult(isin, description, ok=True,
                                   symbol=info.get("symbol", ""),
                                   exchange=info.get("exchange", ""),
                                   currency=cur_label, source="investing.com",
                                   prices=converted)
            inv_reason = (f"investing.com: trovato {info.get('symbol')} in "
                          f"{info.get('currency')} ma cambio non recuperabile")
    except Exception as exc:
        inv_reason = f"investing.com: errore imprevisto ({exc})"

    fd_reason = ""
    try:
        from fondidoc_fetcher import fetch_from_fondidoc
        df, info, fd_reason = fetch_from_fondidoc(isin, start, end)
        if not df.empty:
            return FetchResult(isin, description, ok=True,
                               symbol=info["symbol"], exchange=info["exchange"],
                               currency="EUR", source="fondidoc.it", prices=df)
    except Exception as exc:
        fd_reason = f"fondidoc: errore imprevisto ({exc})"

    res = _fetch_from_yahoo(isin, description, start, end, pause)
    if not res.ok:
        res.reason = f"{inv_reason} || {fd_reason} || {res.reason}"
    return res


def _suffix_rank(symbol: str) -> int:
    for i, suf in enumerate(SUFFIX_PREFERENCE):
        if symbol.upper().endswith(suf):
            return i
    return len(SUFFIX_PREFERENCE)


def _search_candidates(isin: str) -> list[dict]:
    """Cerca l'ISIN su Yahoo Finance e restituisce i candidati ordinati."""
    quotes = []
    try:
        s = yf.Search(isin, max_results=15)
        quotes = s.quotes or []
    except Exception:
        quotes = []
    if not quotes:
        quotes = [{"symbol": isin, "exchDisp": "?", "quoteType": "EQUITY"}]
    cands = [
        q for q in quotes
        if q.get("symbol") and q.get("quoteType", "EQUITY").upper() in ACCEPTED_TYPES
    ]
    cands.sort(key=lambda q: _suffix_rank(q["symbol"]))
    return cands


def _yahoo_symbol_variants(candidates: list[dict]) -> list[dict]:
    """
    Espande i candidati con varianti di borsa: la ricerca per ISIN spesso
    restituisce solo il listing di Londra anche quando esiste il listing
    EUR a Milano o su Xetra con lo stesso ticker base.
    """
    out, seen = [], set()
    for q in candidates:
        sym = q["symbol"]
        base = sym.split(".")[0]
        variants = [sym] + [base + suf for suf in (".MI", ".DE", ".AS", ".PA", ".F")]
        for v in variants:
            if v not in seen:
                seen.add(v)
                out.append({**q, "symbol": v})
    out.sort(key=lambda q: _suffix_rank(q["symbol"]))
    return out


def _download_weekly_eur(symbol: str, start, end) -> tuple[pd.DataFrame, str]:
    """
    Scarica i prezzi settimanali per un simbolo Yahoo.
    Ritorna (DataFrame Data/Prezzo, valuta). Vuoto se niente dati.
    Nota: le barre settimanali Yahoo sono etichettate col lunedi'.
    """
    t = yf.Ticker(symbol)
    hist = t.history(start=start, end=end, interval="1wk", auto_adjust=False)
    if hist is None or hist.empty or "Close" not in hist.columns:
        return pd.DataFrame(), ""
    currency = (t.history_metadata or {}).get("currency", "") or ""
    df = hist.reset_index()[["Date", "Close"]].dropna()
    df["Date"] = pd.to_datetime(df["Date"]).dt.tz_localize(None).dt.normalize()
    df = df.rename(columns={"Date": "Data", "Close": "Prezzo"})
    return df, currency


def _fetch_from_yahoo(isin: str, description: str, start, end,
                      pause: float = 0.6) -> FetchResult:
    """Recupero da Yahoo Finance (fallback). Preferenza EUR, poi conversione."""
    candidates = _yahoo_symbol_variants(_search_candidates(isin))
    if not candidates:
        return FetchResult(isin, description, ok=False,
                           reason="Yahoo: ISIN non trovato")

    tried = []
    non_eur = None  # primo risultato valido in altra valuta
    for q in candidates:
        symbol = q["symbol"]
        try:
            df, currency = _download_weekly_eur(symbol, start, end)
        except Exception as exc:
            tried.append(f"{symbol}: errore ({exc})")
            continue
        finally:
            time.sleep(pause)

        if df.empty:
            tried.append(f"{symbol}: nessun dato nel periodo")
            continue
        if currency.upper() != "EUR":
            tried.append(f"{symbol}: valuta {currency or 'sconosciuta'}")
            if non_eur is None and currency:
                non_eur = (q, df, currency)
            continue

        return FetchResult(isin, description, ok=True, symbol=symbol,
                           exchange=q.get("exchDisp", ""), currency="EUR",
                           source="Yahoo Finance", prices=df)

    if non_eur is not None:
        q, df, currency = non_eur
        converted, cur_label = _to_eur(df, currency, start, end)
        if not converted.empty:
            return FetchResult(isin, description, ok=True, symbol=q["symbol"],
                               exchange=q.get("exchDisp", ""), currency=cur_label,
                               source="Yahoo Finance", prices=converted)
        tried.append(f"{q['symbol']}: cambio {currency}→EUR non recuperabile")

    return FetchResult(isin, description, ok=False,
                       reason="Yahoo: nessuna quotazione utilizzabile. Tentativi: "
                              + "; ".join(tried[:6]))


def read_asset_list(file) -> pd.DataFrame:
    """Legge il file Excel della lista asset (colonne ISIN, Descrizione Asset)."""
    df = pd.read_excel(file)
    df.columns = [str(c).strip() for c in df.columns]
    required = {"ISIN", "Descrizione Asset"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            "Colonne mancanti nel file: " + ", ".join(sorted(missing))
            + ". Attese: ISIN, Descrizione Asset."
        )
    df = df[["ISIN", "Descrizione Asset"]].dropna(subset=["ISIN"])
    df["ISIN"] = df["ISIN"].astype(str).str.strip()
    df["Descrizione Asset"] = df["Descrizione Asset"].astype(str).str.strip()
    return df.drop_duplicates(subset=["ISIN"]).reset_index(drop=True)


def build_output_workbook(results: list[FetchResult]) -> bytes:
    """
    File xlsx in formato TemplatePrezzi: ISIN, Descrizione Asset,
    Data (dd/mm/aaaa), Prezzo Corrente (EUR). Secondo foglio 'Non Recuperati'
    se ci sono asset scartati.
    """
    import io
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ISIN", "Descrizione Asset", "Data", "Prezzo Corrente (EUR)"])

    for r in results:
        if not r.ok:
            continue
        for _, row in r.prices.iterrows():
            ws.append([r.isin, r.description, row["Data"].to_pydatetime(),
                       round(float(row["Prezzo"]), 6)])
            ws.cell(row=ws.max_row, column=3).number_format = "DD/MM/YYYY"

    for col, width in zip("ABCD", (16, 32, 14, 22)):
        ws.column_dimensions[col].width = width

    skipped = [r for r in results if not r.ok]
    if skipped:
        ws2 = wb.create_sheet("Non Recuperati")
        ws2.append(["ISIN", "Descrizione Asset", "Motivo"])
        for r in skipped:
            ws2.append([r.isin, r.description, r.reason])
        for col, width in zip("ABC", (16, 32, 90)):
            ws2.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
